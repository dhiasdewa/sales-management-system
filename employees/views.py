from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from .models import Employee
from .models import Product
from django.db.models import Sum
from .models import Product, Purchase, Sale
from .models import ActivityLog
from .decorators import role_required
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Sale
import json
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.utils.dateparse import parse_date
from .forms import SaleForm
from django.shortcuts import redirect
from django.utils import timezone
from datetime import datetime



def custom_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            if user.is_superuser:
                return redirect("/admin/")

            role_name = user.role.name.lower()

            if role_name == "hr":
                return redirect("hr_dashboard")

            if role_name == "employees":   # SESUAIKAN DENGAN DATABASE KAMU
                return redirect("employee_dashboard")

        else:
            messages.error(request, "Invalid credentials")

    return render(request, "login.html")

@login_required
@role_required(["admin", "manager", "hr"])
def create_sale(request):
    if request.method == "POST":
        product = Product.objects.get(id=request.POST["product"])
        quantity = int(request.POST["quantity"])
        total_price = product.price * quantity

        sale = Sale.objects.create(
            product=product,
            quantity=quantity,
            total_price=total_price
        )

        # 🔥 INI BAGIAN LOGGING
        ActivityLog.objects.create(
            user=request.user,
            action=f"Created sale for {product.name} (Qty: {quantity})"
        )

        return redirect("dashboard")

@login_required
@role_required(["admin", "manager", "hr"])
def create_purchase(request):
    if request.method == "POST":
        product = Product.objects.get(id=request.POST["product"])
        quantity = int(request.POST["quantity"])
        total_price = product.price * quantity

        Purchase.objects.create(
            product=product,
            quantity=quantity,
            total_price=total_price
        )

        # 🔥 LOGGING
        ActivityLog.objects.create(
            user=request.user,
            action=f"Created purchase for {product.name} (Qty: {quantity})"
        )

        return redirect("dashboard")

@login_required
@role_required(["admin", "manager", "hr"])
def create_product(request):
    if request.method == "POST":
        name = request.POST["name"]
        stock = request.POST["stock"]
        price = request.POST["price"]

        product = Product.objects.create(
            name=name,
            stock=stock,
            price=price
        )

        # 🔥 LOGGING
        ActivityLog.objects.create(
            user=request.user,
            action=f"Added new product: {product.name}"
        )

        return redirect("dashboard")


@login_required
@role_required(["Employees", "Admin"])
def employee_dashboard(request):

    user = request.user
    employee = Employee.objects.get(user=user)

    sales = Sale.objects.filter(employee=user)

    total_sales = sales.aggregate(
        total_quantity=Sum("quantity"),
        total_revenue=Sum("total_price")
    )

    now = timezone.now()
    current_month_sales = Sale.objects.filter(
        employee=user,
        date__year=now.year,
        date__month=now.month
    )

    monthly_revenue = current_month_sales.aggregate(
        total=Sum("total_price")
    )["total"] or 0

    # Hitung progress %
    if employee.monthly_target > 0:
        progress = (monthly_revenue / employee.monthly_target) * 100
    else:
        progress = 0



    context = {
        "employee": employee,  # sekarang ini model Employee
        "sales": sales,
        "total_quantity": total_sales["total_quantity"] or 0,
        "total_revenue": total_sales["total_revenue"] or 0,
        "monthly_revenue": monthly_revenue,
        "progress": round(progress, 2),
    }

    return render(request, "employee_dashboard.html", context)


@login_required
@role_required(["admin", "manager", "HR"])
def hr_dashboard(request):

    products = Product.objects.all()
    employees = Employee.objects.all()
    total_products = Product.objects.count()
    total_sales = Sale.objects.count()
    total_purchase = Purchase.objects.count()

    total_revenue = Sale.objects.aggregate(
        total=Sum("total_price")
    )["total"] or 0

    low_stock = Product.objects.filter(stock__lt=5)

    # 🔥 WAJIB ADA INI
    logs = ActivityLog.objects.all().order_by("-timestamp")[:10]

    # SALES PER DAY
    sales_by_day = (
        Sale.objects
        .annotate(sale_date=TruncDate("date"))
        .values("date")
        .annotate(total_sales=Sum("quantity"))
        .order_by("date")
    )

    dates = [str(item["date"]) for item in sales_by_day]
    sales_data = [item["total_sales"] for item in sales_by_day]

    sales = Sale.objects.all()

    # TOP 3 PRODUCTS
    top_products = (
        Sale.objects
        .values("product__name")
        .annotate(total_quantity=Sum("quantity"))
        .order_by("-total_quantity")[:3]
    )

    context = {
        "sales": sales,
        "top_products": top_products,
    }

# REVENUE PER DAY
    revenue_by_day = (
        Sale.objects
        .annotate(sale_date=TruncDate("date"))
        .values("date")
        .annotate(total_revenue=Sum("total_price"))
        .order_by("date")
    )

    revenue_data = [float(item["total_revenue"]) for item in revenue_by_day]


    context = {
        "employees": employees,
        "total_products": total_products,
        "total_sales": total_sales,
        "total_purchase": total_purchase,
        "total_revenue": total_revenue,
        "low_stock": low_stock,
        "logs": logs,  # 🔥 ini baru aman
        "chart_dates": json.dumps(dates),
        "sales_chart_data": json.dumps(sales_data),
        "revenue_chart_data": json.dumps(revenue_data),

    }

    return render(request, "hr_dashboard.html", context)

@role_required(["Employee"])
def add_sale(request):
    if request.method == "POST":
        form = SaleForm(request.POST)
        if form.is_valid():
            sale = form.save(commit=False)

            sale.employee = request.user
            sale.total_price = sale.product.price * sale.quantity

            # Kurangi stock
            sale.product.stock -= sale.quantity
            sale.product.save()

            sale.save()
            return redirect("employee_dashboard")
    else:
        form = SaleForm()

    return render(request, "add_sale.html", {"form": form})

def export_sales_excel(request):

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    sales = Sale.objects.all()

    # FILTER DATE RANGE
    if start_date and end_date:
        sales = sales.filter(date__date__range=[start_date, end_date])

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    headers = ["Date", "Product", "Quantity", "Total Price"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_revenue = 0
    row_number = 2

    for sale in sales:
        sale_date = sale.date.replace(tzinfo=None)

        ws.append([
            sale_date,
            sale.product.name,
            sale.quantity,
            sale.total_price,
        ])

        ws.cell(row=row_number, column=4).number_format = '"Rp" #,##0'

        total_revenue += sale.total_price
        row_number += 1

    ws.append(["", "", "TOTAL", total_revenue])

    total_row = ws.max_row
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=4).number_format = '"Rp" #,##0'

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=sales_report.xlsx'

    wb.save(response)
    return response


def export_top_selling_products(request):

    # GROUP BY PRODUCT
    top_products = (
        Sale.objects
        .values("product__name")
        .annotate(
            total_quantity=Sum("quantity"),
            total_revenue=Sum("total_price")
        )
        .order_by("-total_quantity")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Top Selling Products"

    headers = ["Product", "Total Quantity Sold", "Total Revenue"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_number = 2

    for item in top_products:
        ws.append([
            item["product__name"],
            item["total_quantity"],
            item["total_revenue"],
        ])

        ws.cell(row=row_number, column=3).number_format = '"Rp" #,##0'
        row_number += 1

    # Auto column width
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=top_selling_products.xlsx'

    wb.save(response)
    return response


@login_required
def user_logout(request):
    logout(request)
    return redirect("login")

@login_required
def inventory_list(request):
    products = Product.objects.all()
    return render(request, "inventory.html", {"products": products})


